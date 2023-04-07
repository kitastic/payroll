import csv
import openpyxl
import os.path
from selenium import webdriver
# no longer need to download browser drivers
import chromedriver_autoinstaller
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import time
import re
import glob                     # check if file exists with wildcard
import datetime
import PySimpleGUI as sg

# local files
import helper
import xlHelper
from pathlib import Path

class WebBot:

    def __init__(self, bundle):
        self.username = bundle[0]
        self.password = bundle[1]
        self.activeEmp = bundle[2]
        self.workBook = bundle[3]
        # start web auto
        chromedriver_autoinstaller.install()
        opts = webdriver.ChromeOptions()


        dlDir = r'{}\{}\\'.format(os.getcwd(),'dl')
        # make sure path exists
        Path(dlDir).mkdir(parents=True,exist_ok=True)
        prefs = {'download.default_directory': dlDir,
                 'directory_upgrade': True,
                 }
        opts.add_experimental_option('prefs', prefs)
        self.driver = webdriver.Chrome(options=opts)
        self.originalData = ''  # text derived from webelement
        self.rawData = []  # data divided into employee lists
        self.parsedData = []  # rawData further divided into list of 4 cols
        self.sheet = ''
        self.income = []
        self.startDate = ''

    def login(self):
        self.driver.get("https://pos2.zota.us/#/login")
        # self.driver.maximize_window()
        usernameInput = self.driver.find_element(By.NAME, "username")
        usernameInput.clear()
        usernameInput.send_keys(self.username)
        passInput = self.driver.find_element(By.NAME, "password")
        passInput.clear()
        passInput.send_keys(self.password)
        loginBtn = self.driver.find_element(By.CLASS_NAME, "btn-login")
        loginBtn.click()
        # once logged in, wait for page to load (when avatar is located)
        helper.waitLoadingPresence('avatar', 10, self.driver)

    def getPayroll(self, startDay, endDay):
        # save sheet name for when exporting to xlsx
        self.sheet = startDay
        # load payroll page
        self.driver.get("https://pos2.zota.us/#/payroll")

        # wait for startdate element to be clickable
        s = '//div/div/div[2]/div/main/div/div/div/div[3]/div[1]/div/div[1]/div/div/div/div[2]/span/span/span/span/input'
        helper.waitLoadingClick(s, 5, 'By.XPATH', self.driver)

        # select start date
        startDate = self.driver.find_element(By.NAME, 'startDate')
        defaultstart = helper.parseDate(startDate.get_attribute('value'))
        start = helper.parseDate(startDay)
        helper.enterDate(start, defaultstart, startDate)
        time.sleep(.5)

        # select and change end date
        endDate = self.driver.find_element(By.NAME, 'endDate')
        defaultend = helper.parseDate(endDate.get_attribute('value'))
        end = helper.parseDate(endDay)
        helper.enterDate(end, defaultend, endDate)
        time.sleep(1)

        # manually clicking menu to select employee from list
        e = self.driver.find_element(By.CLASS_NAME, 'k-dropdown-wrap')
        for a in self.activeEmp:
            e.click()
            end = False
            eZota = ""
            while (not end):
                # while zota employee list not ended and not matched
                e.send_keys(Keys.ARROW_DOWN)
                etemp = e.text
                if eZota == etemp or a in e.text:
                    end = True
                eZota = etemp

            load = self.driver.find_element(By.XPATH,
                                            '//*[@id="root"]/div/div[2]/div/main/div/div/div/div[3]/div[1]/div/div[4]/button')
            load.click()  # click load payroll button
            reportBtn = '//div/div/div[2]/div/main/div/div/div/div[3]/div[2]/div/div/div/div/div/div/div[2]/div/div/div[1]/div[2]/button'
            helper.waitLoadingClick(reportBtn, 8, "By.XPATH",
                                    self.driver)  # wait for report to load before selecting type of report
            reportBtn = self.driver.find_element(By.XPATH, reportBtn)
            reportBtn.click()
            r = self.driver.find_element(By.XPATH,
                                         '//div/div/div[2]/div/main/div/div/div/div[3]/div[2]/div/div/div/div/div/div/div[1]/div/div/div/pre')
            '''
            Report format is multiple lines of strings. Each line will now be split at \n character
            and then added to a list for each employee. Then the full employee list will be added
            to a bigger list containing per employee.
            '''
            self.rawData.append(list(r.text.split('\n')))  # report is multiple lines of strings
            self.originalData = self.originalData + r.text + '\n'

    def parseData(self):
        for i in self.rawData:
            # each i is for each employee ticket for the whole week
            # out of the 26 lines of raw data, we will only extract what we need
            dates = re.findall('\d+/\d+/\d+', i[1])
            rDate = dates[0]
            tempParsed = []
            tempIncome = []
            tempParsed.append(['======', '======', '======', '======'])
            tempParsed.append(['Date', '', dates[0], dates[1]])
            tempParsed.append(['Name', '', '', re.search('\w+$', i[2]).group(0)])
            tempParsed.append(['Pay', '-----', '-----', '-----'])
            tempParsed.append(['Total Sale', '', '', helper.getFloat(i[6])])
            tempParsed.append(['', '', '', '', ])
            tempParsed.append(['Commission', '', '', helper.getFloat(i[8])])
            tempParsed.append(['Tips', '', '', helper.getFloat(i[9])])
            tempParsed.append(['', '', '', '-------'])
            tempParsed.append(['Total Pay', '', '', helper.getFloat(i[11])])
            tempParsed.append(['Check', '', '', helper.getFloat(i[12])])
            tempParsed.append(['Cash', '', '', helper.getFloat(i[13])])
            tempParsed.append(['Daily:', '-------', '-------', '-------'])
            tempParsed.append(['Day', 'Total', 'Comm', 'Tips'])
            day, values = helper.getDay(i[17])
            tempParsed.append([day, values[0], values[1], values[2]])
            dateObj, dWord = helper.dayToDate(day, rDate)
            tempIncome.append([dateObj, dWord, values[0], values[1], values[2]])
            day, values = helper.getDay(i[18])
            tempParsed.append([day, values[0], values[1], values[2]])
            dateObj, dWord = helper.dayToDate(day, rDate)
            tempIncome.append([dateObj, dWord, values[0], values[1], values[2]])
            day, values = helper.getDay(i[19])
            tempParsed.append([day, values[0], values[1], values[2]])
            dateObj, dWord = helper.dayToDate(day, rDate)
            tempIncome.append([dateObj, dWord, values[0], values[1], values[2]])
            day, values = helper.getDay(i[20])
            tempParsed.append([day, values[0], values[1], values[2]])
            dateObj, dWord = helper.dayToDate(day, rDate)
            tempIncome.append([dateObj, dWord, values[0], values[1], values[2]])
            day, values = helper.getDay(i[21])
            tempParsed.append([day, values[0], values[1], values[2]])
            dateObj, dWord = helper.dayToDate(day, rDate)
            tempIncome.append([dateObj, dWord, values[0], values[1], values[2]])
            day, values = helper.getDay(i[22])
            tempParsed.append([day, values[0], values[1], values[2]])
            dateObj, dWord = helper.dayToDate(day, rDate)
            tempIncome.append([dateObj, dWord, values[0], values[1], values[2]])
            day, values = helper.getDay(i[23])
            tempParsed.append([day, values[0], values[1], values[2]])
            dateObj, dWord = helper.dayToDate(day, rDate)
            tempIncome.append([dateObj, dWord, values[0], values[1], values[2]])
            tempParsed.append(['     ', '-------', '-------', '-------'])
            values = helper.getDay(i[25])
            tempParsed.append(['', values[0], values[1], values[2]])
            self.parsedData.append(tempParsed.copy())
            self.income.append(tempIncome.copy())
            tempParsed.clear()
            tempIncome.clear()

    def dlEmpSales(self, startDate, endDate, salonName,):

        self.startDate = startDate
        saveAsName = salonName[0] + '.' + startDate.replace('/', '.') + '.xlsx'
        self.driver.get('https://pos2.zota.us/#/reports/employee-reports')
        sDateXpath = '//div/div/div[2]/div/main/div/div/div[3]/div[2]/span/span/span/span/input'
        helper.waitLoadingClick(sDateXpath, 10, 'By.XPATH: startDate', self.driver)

        sDateElement = self.driver.find_element(By.NAME, 'startDate')
        defaultStart = helper.parseDate(sDateElement.get_attribute('value'))
        sDateFormatted = helper.parseDate(self.startDate)
        helper.enterDate(sDateFormatted, defaultStart, sDateElement)
        time.sleep(.5)

        eDateElement = self.driver.find_element(By.NAME, 'endDate')
        defaultEnd = helper.parseDate(eDateElement.get_attribute('value'))
        eDateFormatted = helper.parseDate(endDate)
        helper.enterDate(eDateFormatted, defaultEnd, eDateElement)
        time.sleep(.5)

        runReportBtn = self.driver.find_element(By.XPATH, '//*[@id="root"]/div/div[2]/div/main/div/div/div[2]/div[5]/button')
        runReportBtn.click()

        # wait to locate and click the excel download button
        excelxpath = '//div/div/div[2]/div/main/div/div/div[4]/div/div/div/div/div/div/div/div[2]/span[2]'
        helper.waitLoadingClick(excelxpath, 10, 'By.XPATH: excel link', self.driver)
        try:
            self.driver.find_element(By.XPATH, excelxpath).click()
            sg.easy_print('Successfuly clicked download link')
        except Exception as e:
            print('Failed to click excel link')



    def exportParToXlsx(self, fname):
        # hard-coded date for now and changed / to . because of xlsx error
        self.sheet = self.parsedData[0][1][3].replace('/', '.')
        # check to see if .xlsx exists or create one
        if os.path.isfile(fname):
            # if book exists, now check for sheet exists
            book = openpyxl.load_workbook(fname,)  # load book
            if self.sheet in book.sheetnames:
                print('Sheet name already exists: OVERWRITING')

            # create a new one by setting sheet name as active, this deletes existing sheet
            self.sheet = book.active
            for employee in self.parsedData:
                for lines in employee:
                    self.sheet.append(lines)
            book.save(fname)
        else:
            book = openpyxl.Workbook()  # create book
            worksheet = book.create_sheet(self.sheet)
            for employee in self.parsedData:
                for lines in employee:
                    worksheet.append(lines)
            book.save(fname)

    def parseXlToData(self, sheetName):
        workBook = openpyxl.load_workbook(self.workBook)
        workSheet = workBook[sheetName]
        dates = re.findall('\d+/\d+/\d+', workSheet['A3'].value)
        firstLine = ['Date', '', dates[0], ' - ' + dates[1]]

        techTable = []
        for tech in self.activeEmp:
            techTable.append(firstLine)
            techTable.append(['Name', '', '', tech])
            techTable.append(['Pay', '-----', '-----', '-----'])
            totalSale = 0
            commission = 0
            tips = 0
            totalPay = 0
            check = 0
            cash = 0
            tempDailies = []
            tempIncome = []
            # search through all rows in sheet to find current employee's
            # daily incomes
            day = ''
            for row in workSheet.iter_rows(values_only=True):

                if isinstance(row[0], datetime.datetime):
                    day = row[0].strftime('%d-%a')
                elif not row[0]:
                    continue
                else:
                    try:
                        if tech in row[0]:
                            if row[1] == 'S':
                                totalSale = totalSale + row[4]
                                commission = commission + row[11]
                                tips = tips + row[10]
                                tempDailies.append([day,row[4],row[11],row[10]])
                    except Exception as e:
                        print('Cannot iterate to find tech')

            techTable.append(['Total Sale', '', '', totalSale])
            techTable.append(['', '', '', '', ])
            techTable.append(['Commission', '', '', commission])
            techTable.append(['Tips', '', '', tips])
            techTable.append(['', '', '', '-------'])
            techTable.append(['Total Pay', '', '', commission+tips])
            techTable.append(['Check', '', '', (commission * .6) + tips])
            techTable.append(['Cash', '', '', commission * .4])
            techTable.append(['Daily:', '-------', '-------', '-------'])
            techTable.append(['Day', 'Total', 'Comm', 'Tips'])
            for row in tempDailies:
                techTable.append(row.copy())
                tempIncome.append(row.copy())
            techTable.append(['     ', '-------', '-------', '-------'])
            techTable.append(['', totalSale, commission, tips])
            self.income.append(tempIncome.copy())
            self.parsedData.append(techTable.copy())
            tempIncome.clear()
            techTable.clear()

    def printOriginalData(self):
        print(self.originalData)

    def exportOrigToTxt(self, filename):
        with open(filename, 'w') as f:
            f.writelines(self.originalData)

    def importTxtToOrig(self, filename):
        with open(filename, 'r') as f:
            self.originalData = f.readlines()

    def origToRaw(self):
        # this function is for working with data already saved without loading website
        # rawData is a list of lists, the sublists are divided by employee
        marker = 0
        for emp in range(len(self.activeEmp)):
            temp = []  # list per employee
            for i in range(26):
                # length of data per employee is 26, but index is 25
                temp.append(self.originalData[marker].replace('\n', ''))
                marker = marker + 1
            self.rawData.append(temp.copy())
            temp.clear()

    def exportParToCsv(self, fname):
        with open(fname, 'w', newline='') as f:
            writer = csv.writer(f)
            for employee in self.parsedData:
                for lines in employee:
                    writer.writerows(lines)


    def exportEmployee(self, name):
        index = self.activeEmp.index(name)
        return [self.parsedData[index].copy(), self.income[index].copy()]

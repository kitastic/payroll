import os
import openpyxl
import datetime
import pandas as pd
import re
import xlHelper
import json
import Bot
import PySimpleGUI as sg
import Employee
from pprint import PrettyPrinter
import num2words
import string
from pathlib import Path

pp = PrettyPrinter(
    indent=2,
    width=100,
    compact=True,
    sort_dicts=False,
)


class Salon(Bot.Bot):
    def __init__(self,bundle):
        """

        Args:
            bundle: (dictionary) layout:
                    bundle =  { 'name': 'upscale',
                                'login': {'username': 'upscalemanager', 'password': 'Joeblack334$'},
                                'salesJson':
                                'salesXl':    # might not need yet
                                'payments': 'uPayments.xlsx',
                                'employees': {
                                name:{'active':True,
                                      'id': idNum, 'name': nameCapitalized, 'salonName':salon, 'pay':pay,
                                      'fees':fees, 'rent':rent, 'printchecks': printchecks,
                                      'paygrade':{'regType':regType, 'cashType': cashtype,
                                                  'janitorType': janitortype, 'checkdealType': checkdealtype, 'owner': owner,
                                      'regular':{'commission': commission, 'check': check},
                                      'special':{'commissionspecial': comspec, 'checkdeal': checkdeal,
                                                 'checkoriginal': checkoriginal, 'cashrate': cashrate}
                                      }}}
        """
        super().__init__()
        # at this point, the keys in bundle have been alphabetized by json
        self.employees = bundle['employees']
        self.zotaUname = bundle['login']['username']
        self.zotaPass = bundle['login']['password']
        self.salonName = bundle['name']
        self.paymentsFnames = bundle['paymentsFnames']
        self.path = bundle['path']
        self.salesFnames = bundle['salesFnames']    # dict of json files names, key=year
        self.startCheckNum = bundle['startchecknum']

        # variables stored during program runtime
        self.salesDict = dict()  # {year: {datetime: {empName: [total, comm, tips]}}
        self.Emps = dict()  # holds employee objects
        self.setupSalon()

    def setupSalon(self):
        """
            Retrieves all sales data from json and creates employee objects
        Returns:
            None
        """
        if self.salesFnames:
            for year, fname in self.salesFnames.items():
                try:
                    if os.path.getsize(self.path+fname) > 10:
                        with open(self.path+fname,'r') as reader:
                            self.salesDict = json.load(reader)
                except FileNotFoundError:
                    print(f'[Salon.setupSalon] {self.salonName} did not find any json')

        for emp,info in self.employees.items():
            if info['paygrade']['cashType']:
                self.Emps[string.capwords(emp)] = Employee.EmployeeCash(info)
            elif info['paygrade']['checkdealType']:
                self.Emps[string.capwords(emp)] = Employee.EmployeeSpecial(info)
            elif info['paygrade']['janitorType']:
                self.Emps[string.capwords(emp)] = Employee.EmployeeJanitor(info)
            else:
                self.Emps[string.capwords(emp)] = Employee.Employee(info)

    def createEmpFromGui(self,name,empData):
        self.Emps[name] = Employee.Employee(empData[name])

    def createEmpReg(self,name):
        newEmp = {
            name:{'active':True,
                  'id':0,'name':name,'salonName':self.salonName,'pay':0,'fees':0,'rent':0,
                  'printchecks': True,
                  'paygrade':{'regType':True,
                              'cashType': False,
                              'janitorType': False,
                              'checkdealType': False,
                              'owner': False,
                              'regular':{'commission':6,'check':6},
                              'special':{'commissionspecial':0,'checkdeal':0,
                                         'checkoriginal':0, 'cashrate': 0}
                              }}}
        self.Emps[name] = Employee.Employee(newEmp[name])

    def deleteEmp(self,name):
        deletedValue = self.Emps.pop(name)
        print(deletedValue)

    def exportPayroll(self, sDate, format):
        empdata = {}
        xldict = dict()
        for name,obj in self.Emps.items():
            printableData = ''
            try:
                printableData = obj.getPrintOut()
            except Exception:
                sg.popup_ok('ERROR: Salon.exportPayroll: employee data is empty\n'
                            'Try to calculate payroll first')
            if len(printableData) > 5:
                if not obj.janitortype and not obj.owner:
                    empdata[name] = printableData
                if format == 'txt':
                    # make sure path exists
                    Path(f'../payroll/{self.salonName}/').mkdir(parents=True,exist_ok=True)
                    pfname = f'../payroll/{self.salonName}/{self.salonName[0]}.{sDate.replace("/",".")}.{name}.txt'
                    Path(pfname.replace(f'{self.salonName[0]}.{sDate.replace("/",".")}.{name}.txt','')).mkdir(parents=True,
                                                                                                              exist_ok=True)
                    with open(pfname,'w') as f:
                        f.writelines(printableData)
                    print(f'INFO: ({self.salonName}) finished exporting text files')
        if format == 'html':
            htmlheader = """
                        <!DOCTYPE html>
                        <html>
                        <head>
                        <style>
                        @media print {
                            .pagebreak {
                                clear: both;
                                page-break-before: always;
                            }
                        }
                        </style>
                        </head>
                        <pre>
                        <body>"""
            htmlpagebreak = """<div class="pagebreak"></div>"""
            htmlfooter = """
            </body>
            <pre>
            </html>
            """
            pfname = f'../payroll/{sDate.replace("/",".")}.{self.salonName}.html'
            with open(pfname, 'w+') as write:
                write.writelines(htmlheader)
                write.write('\n')
                lastValue = list(empdata.values())[-1]
                for values in empdata.values():
                    write.writelines(values)
                    if values != lastValue:
                        write.writelines(htmlpagebreak)
                write.writelines(htmlfooter)
            print(f'[Salon.exportPayroll]: {self.salonName} finished exporting {pfname}')
            # self.printHtml(pfname)

            # now update yearly excel book for 1099
            # pass path, salon prefix for sheetname, data
            path = f'../payroll/{sDate[-4:]}.xlsx'
            sheet = f'{self.salonName[0].lower()}.salary'
            data = []
            for emp, obj in self.Emps.items():
                if obj.printchecks or obj.cashtype:
                    xldict = xldict | {emp: {}}
                    # remove nickname in parentheses
                    xldict[emp] = obj.getXlReport()
                    sdate = datetime.datetime.strptime(sDate, '%m/%d/%Y')
                    edate = sdate + datetime.timedelta(days=6)
                    eDate = datetime.datetime.strftime(edate, '%m/%d/%Y')
                    xldict[emp]['name'] = re.search('^\w+ \w+', emp).group(0)
                    xldict[emp]['date'] = eDate
                    xldict[emp]['memo'] = f'{sDate} - {eDate} PAYROLL'
                    data.append([sDate, eDate, emp.upper(), xldict[emp]['cash'], xldict[emp]['check'], xldict[emp]['checkdeal']])
            df = pd.DataFrame(data,)
            reader = pd.read_excel(path, sheet_name=sheet, index_col=False)
            startRow = len(reader.index) + 1
            with pd.ExcelWriter(path, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
                df.to_excel(writer, sheet_name=sheet, header=False, index=False, startrow=startRow)

        # create check printouts
        # ----------------------------------------------
        # first copy template checks from excel sheet and create new sheet
        # xlHelper preserves column and row sizes which is important
        # TODO catch error if sheet does not exist create one when opening
        # sourcepfname = f'../payroll/checksTemplates.xlsx'
        # sourcesheet = f'{self.salonName[0].lower()}.check'
        # pfname = f'../payroll/{self.salonName[0].lower()}.checks.xlsx'
        # targetsheet = f'{sDate[0:5].replace("/", ".")}'
        # wb_target = openpyxl.Workbook()
        # target_sheet = wb_target.create_sheet(targetsheet)
        # wb_source = openpyxl.load_workbook(sourcepfname,data_only=True)
        # source_sheet = wb_source[sourcesheet]
        # xlHelper.copy_sheet(source_sheet,target_sheet)
        #
        # counter = 0
        # row = {
        #         "checkNum": 2,
        #         "checkDate": 3,
        #         "name": 5,
        #         "amtNum": 5,
        #         "amtWord": 6,
        #         "memo": 8,
        #         "micrCheckNum": 10}
        # nextCheckIndex = 11
        # nextPageCheckIndex = 23
        # for emp, values in xldict.items():
        #     if values['check'] == 0:
        #         pass
        #     else:
        #         target_sheet[f'J{row["checkNum"]}'] = self.startCheckNum
        #         target_sheet[f'i{row["checkDate"]}'] = datetime.date.today().strftime('%m/%d/%Y')
        #         target_sheet[f'd{row["name"]}'] = values['name']
        #         target_sheet[f'i{row["amtNum"]}'] = f'{values["check"]:.2f}'
        #         target_sheet[f'c{row["amtWord"]}'] = f'{num2words.num2words(values["check"]).replace(",", "").replace(" and", "").upper()}'
        #         target_sheet[f'd{row["memo"]}'] = values["memo"]
        #         target_sheet[f'g{row["micrCheckNum"]}'] = self.startCheckNum
        #         self.startCheckNum += 1
        #         counter += 1
        #         if counter % 3 == 0:
        #             for key in row:
        #                 row[key] += nextPageCheckIndex
        #         else:
        #             for key in row:
        #                 row[key] += nextCheckIndex
        # wb_target.save(pfname)


    def getDataToSave(self):
        emps = {}
        for empObjKeys,obj in self.Emps.items():
            emps[empObjKeys] = obj.getInfo()
        data = {'name':self.salonName,
                'login':{'username':self.zotaUname,'password':self.zotaPass},
                'salesFnames': self.salesFnames,
                'path': self.path,
                'employees':emps,
                'paymentsFnames': self.paymentsFnames,
                'startchecknum': self.startCheckNum
                }

        return data

    def getEmps(self,name=None):
        emps = {}
        for name,empObj in self.Emps.items():
            emps[name] = empObj.getInfo()
        return emps

    def getEmpStatus(self):
        status = {}
        for name,obj in self.Emps.items():
            result = obj.getStatus()
            if result:
                status[name] = result
        return status

    def getJsonRange(self,sDate,eDate):
        """
            With the given date range, returns a dictionary of days with their
            respective income for all employees.
            have to load json file, sometimes current and previous commands
            have not populated the internal json; ie this is the first command
            when starting up the gui window. When the range spans two years,
            just gather data from 2 jsons into one temp dict and then iterate
            for given range
        Args:
            sDate: (string)
            eDate: (string)
        Returns:
            Dictionary where keys are datetime and values are dictionaries of
            employees and their income
        """
        sdate = datetime.datetime.strptime(sDate,'%m/%d/%Y')
        edate = datetime.datetime.strptime(eDate,'%m/%d/%Y')
        range = []
        if sdate.year != edate.year:
            start = sdate.year
            while start <= edate.year:
                range.append(start)
                start += 1
        else:
            range = [sdate.year]

        tmpSales = dict()
        # grab all dates from, maybe both years into tmpSales
        for year in range:
            if not os.path.isfile(f'{self.path}{self.salonName}Sales{year}.json'):
                print(f'[Salon.getJsonRange] {self.salonName} file not found')
            else:
                with open(f'{self.path}{self.salonName}Sales{year}.json', 'r') as reader:
                    tmpSales = tmpSales | json.load(reader)

        if not tmpSales:
            # if there is not any file for any years in range
            return False

        keys = [datetime.datetime.strptime(i,'%m/%d/%Y') for i in tmpSales]
        wantedRange = dict()
        for k in keys:
            if k >= sdate and k <= edate:
                # convert key back to string to match json
                kstr = datetime.datetime.strftime(k,'%m/%d/%Y')
                wantedRange[k] = tmpSales[kstr]
        return wantedRange

    def getPayroll(self, sDate, eDate):
        """
            given startdate and enddate, salon will tell each employee to calculate
            their own payroll and return their report back
        Args:
            request: (list) [startDate, endDate]

        Returns:
            dictionary of employee key and their payroll values
        """
        week = self.getJsonRange(sDate,eDate)
        if not week:
            return False

        sorted = {}
        # rearrange dictionary keys from days to employees
        for dates,value in week.items():
            for e in value:
                sorted[e.lower()] = {}
        for dates,value in week.items():
            for e,total in value.items():
                sorted[e.lower()][dates] = total
        # pp.pprint(sorted)
        # compare for extra employees , ie 'anybody*', not currently in settings DB and create new regular ones
        salesEmployees = [string.capwords(n) for n in sorted]
        currentEmployees = []
        for key,obj in self.Emps.items():
            currentEmployees.append(string.capwords(key))
        # create Employees for any extra in sales so they can calculate their sales
        for e in salesEmployees:
            if e not in currentEmployees:
                agree = sg.popup_ok_cancel(
                    f'{e} not found in current employees list.\n\nWould you like to add employee to database?')
                if agree == 'OK':
                    self.createEmpReg(e)
                else:
                    print(f'INFO: skipping payroll calculations for {e}.')
        # now tell all employees to calculate
        payrollPkt = {}
        for eName,eObj in self.Emps.items():
            for e,val in sorted.items():
                if eName in string.capwords(e):
                    eObj.calculatePayroll(val)
                    payrollPkt[eName] = eObj.getPrintOut()
            if eObj.janitortype:
                eObj.calculatePayroll(sales=None)
                payrollPkt[eObj.name] = eObj.getPrintOut()
        return payrollPkt

    def getSalonInfo(self):
        """
            Used for displaying salon info in the salon tab when clicking the load salon button
        Returns:
            list of dictionary login and formatted additional text dislay to go into the output in salon tab
        """
        login = {
            'username': self.zotaUname,
            'password': self.zotaPass,
        }
        niceprint = f'path: {self.path},\nsalesFnames: {self.salesFnames}' \
                    f'\npymentsFnames: {self.paymentsFnames}\n' \
                    f'startingcheckNum: {self.startCheckNum}\n' \
                    f'employees:\n'
        for names in self.Emps:
            niceprint += f'  {names}\n'
        return login, niceprint, self.startCheckNum

    def mergeSheetToBook(self,fname,path,book):
        '''
        copy sheet from new downloaded book and merge it with a book that keeps track of weekly amounts
        sheet name is the first day of the week
        Args:
            fname:
            path:
            book:

        Returns:

        '''
        tempPathAndFname = path + fname
        sheetName = fname.replace('.xlsx','')

        # import new workbook sheet to existing book and delete new book
        wb_target = openpyxl.load_workbook(book)

        # before copying new temp sheet into weeklyDB, delete existing match
        if sheetName in wb_target.sheetnames:  # remove default sheet
            wb_target.remove(wb_target[sheetName])

        target_sheet = wb_target.create_sheet(sheetName)
        wb_source = openpyxl.load_workbook(tempPathAndFname)
        source_sheet = wb_source['Sales Summary']
        xlHelper.copy_sheet(source_sheet,target_sheet)
        wb_target.save(book)
        # remove temporary downloaded file from zota after extracting info
        os.remove(tempPathAndFname)
    def readSalesXltoJson(self, pfName):
        if not os.path.isfile(pfName):
            print(f'[Salon.readSalesXltoJson]: {self.salonName} cannot read {pfName}')
            return

        workBook = openpyxl.load_workbook(pfName)
        workSheet = workBook.active

        # keys are employee names, values are their totals for the day
        empDict = dict()
        tmpSales = dict()   # daily sales to be inserted in to sales json which is yearly keyed
        day = ''
        currentYr = False

        for row in workSheet.iter_rows(values_only=True):
            if isinstance(row[0],datetime.datetime):
                # if employees dictionary has data, copy it to sales dictionary
                # before setting new day and getting new day data
                if empDict and day != 0:  # if not False
                    tmpSales[day] = empDict.copy()
                    empDict.clear()
                # day is going to be dictionary key, but needs to be
                # converted to string because datetime is not serializable
                # by json
                day = datetime.datetime.strftime(row[0],'%m/%d/%Y')
                year = row[0].year
                if not currentYr:
                    currentYr = year
                else:
                    if year != currentYr:
                        self.salesDict[currentYr] = tmpSales.copy()
                        tmpSales.clear()
                        currentYr = year
            elif not row[0]:
                continue
            else:
                try:
                    if row[1] == 'S':
                        tech = row[0]
                        totalSale = row[4]
                        commission = row[11]
                        tips = row[10]
                        empDict[tech] = [totalSale,commission,tips]
                except Exception as e:
                    print('Cannot iterate to find tech')
        # pack any employee data that still in storage because iterater
        # has not reached a row with datetime
        tmpSales[day] = empDict.copy()
        self.salesDict[currentYr] = tmpSales.copy()
        empDict.clear()
        tmpSales.clear()

    def updateEmpFromGui(self,name,empData):
        # easiest way is to remove existing dictionary and set new one
        self.Emps.pop(name)
        self.Emps[name] = Employee.Employee(empData[name])

    def updateJsonFileDelXl(self,path):
        for year, days in self.salesDict.items():
            fname = f'{self.salonName}Sales{year}.json'
            if os.path.isfile(self.path+fname):
                size = os.path.getsize(self.path+fname)
                if size < 10:
                    # sometimes empty json can have {} and thats 2 bytes
                    # this is an empty file so we can just dump the whole json
                    with open(self.path+fname,'w+') as writer:
                        json.dump(self.salesDict[year],writer,indent=4,sort_keys=True)
                    sg.easy_print('Status: Salon.updateJsonFile > is file but considered empty; overwrite')
                else:
                    with open(self.path+fname,'r') as reader:
                        data = json.load(reader)

                    # find what day ended in file and append dates greater than that
                    # or another way, just jump and do a '|' (straight line up not slash)
                    # to add or up update with the right side taking priority to replace left side
                    self.salesDict[year] = data | self.salesDict[year]

                    with open(self.path+fname,'w+') as writer:
                        json.dump(self.salesDict[year],writer,indent=4,sort_keys=True)
                    print('Status: Salon.updateJsonFile > json exists, merging data')
            else:
                with open(self.path+fname,'w+') as writer:
                    json.dump(self.salesDict[year],writer,indent=4,sort_keys=True)
                    print('Status: Salon.updateJsonFile > no json exists, creating new one')
        if path:
            # now delete recently downloaded excel from website in tmp folder
            filename = max([f for f in os.listdir(path)],
                           key=lambda xa:os.path.getctime(os.path.join(path,xa)))
            os.remove(path + filename)

    def updateSalon(self, salonPkt):
        self.salonName = salonPkt['sname']
        self.zotaUname = salonPkt['login']['username']
        self.zotaPass = salonPkt['login']['password']
        self.startCheckNum = salonPkt['startchecknum']
